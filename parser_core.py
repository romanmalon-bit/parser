import aiohttp
import asyncio
import json
from urllib.parse import urlparse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import random
import datetime
from collections import defaultdict
from pathlib import Path
import math

# =========================
# Глобальні змінні (налаштовуються з проекту)
# =========================
LOCATION = "France"
GL = "fr"
HL = "fr"
BASE_URL = "https://google.serper.dev/search"

TARGET_DOMAINS = set()

RESULTS_PER_PAGE = 10              # скільки результатів просимо за запит
MAX_CONCURRENT_REQUESTS = 3        # як ти попросив

PAGES = 3                          # реальна кількість сторінок (рахується з MAX_POSITIONS)
MAX_POSITIONS = 30                 # Top N

HISTORY_FILE = "serp_history.json"
OUTPUT_FILE = "serp_output.xlsx"

# Динамічні бакети, залежать від MAX_POSITIONS
BUCKET_RANGES = []  # список (start, end)
BUCKET_LABELS = []  # список "start-end"


# =========================
# ПІДТРИМКА СУБДОМЕНІВ
# =========================
def get_full_domain(url: str) -> str:
    try:
        netloc = urlparse(url).netloc.lower()
        if netloc.startswith("www."):
            netloc = netloc[4:]
        elif netloc.startswith("m."):
            netloc = netloc[2:]
        return netloc
    except Exception:
        return ""


def is_target_domain(full_domain: str) -> bool:
    for target in TARGET_DOMAINS:
        if full_domain == target or full_domain.endswith("." + target):
            return True
    return False


# =========================
# Бакети (динамічні діапазони позицій)
# =========================
def build_buckets(max_positions: int):
    """
    Створює:
      BUCKET_RANGES: список (start, end)
      BUCKET_LABELS: список "start-end"

    Логіка:
      1-3, 4-10, 11-20, 21-30, далі 31-40, 41-50, ... до max_positions
    """
    global BUCKET_RANGES, BUCKET_LABELS
    BUCKET_RANGES = []
    BUCKET_LABELS = []

    # Базові бакети
    if max_positions >= 1:
        BUCKET_RANGES.append((1, min(3, max_positions)))
    if max_positions >= 4:
        BUCKET_RANGES.append((4, min(10, max_positions)))
    if max_positions >= 11:
        BUCKET_RANGES.append((11, min(20, max_positions)))
    if max_positions >= 21:
        BUCKET_RANGES.append((21, min(30, max_positions)))

    # Далі — кроками по 10: 31-40, 41-50, ...
    start = 31
    while start <= max_positions:
        end = min(start + 9, max_positions)
        BUCKET_RANGES.append((start, end))
        start += 10

    BUCKET_LABELS = [f"{s}-{e}" for (s, e) in BUCKET_RANGES]


def bucket_for_position(pos: int) -> str:
    """
    Повертає назву бакету для позиції, наприклад "1-3", "4-10", "31-40".
    Якщо позиція > MAX_POSITIONS — повертає ">MAX_POSITIONS".
    """
    for (start, end), label in zip(BUCKET_RANGES, BUCKET_LABELS):
        if start <= pos <= end:
            return label
    return f">{MAX_POSITIONS}"


def calculate_success_score(stats: dict) -> int:
    """
    Рахує "Score" для домену, використовуючи ваги по бакетах:
      1-й бакет: 100
      2-й      : 30
      3-й      : 10
      4-й      : 3
      решта    : 1
    """
    score = 0
    for i, label in enumerate(BUCKET_LABELS):
        if i == 0:
            w = 100
        elif i == 1:
            w = 30
        elif i == 2:
            w = 10
        elif i == 3:
            w = 3
        else:
            w = 1
        score += stats.get(label, 0) * w
    return score


# =========================
# API Key Manager
# =========================
class APIKeyManager:
    def __init__(self, keys):
        self.keys = keys
        self.current_index = 0
        self.lock = asyncio.Lock()
        self.key_fail_counts = defaultdict(int)
        self.max_fails_per_key = 3

    async def get_current_key(self):
        async with self.lock:
            return self.keys[self.current_index]

    async def get_current_index(self):
        async with self.lock:
            return self.current_index

    async def rotate_key(self):
        async with self.lock:
            initial_index = self.current_index
            attempts = 0
            while attempts < len(self.keys):
                self.current_index = (self.current_index + 1) % len(self.keys)
                attempts += 1
                if self.key_fail_counts[self.current_index] < self.max_fails_per_key:
                    print(f"Перехід на ключ #{self.current_index + 1}")
                    return True
                if self.current_index == initial_index:
                    break
            print("Всі ключі вичерпано — чекаємо 60 сек...")
            await asyncio.sleep(60)
            self.key_fail_counts.clear()
            self.current_index = 0
            return True

    async def mark_key_failed(self):
        async with self.lock:
            self.key_fail_counts[self.current_index] += 1
            print(
                f"Помилка ключа #{self.current_index + 1} "
                f"({self.key_fail_counts[self.current_index]}/3)"
            )


# =========================
# Пошук
# =========================
async def serper_search_async(
    query: str,
    page: int,
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    api_key_manager: APIKeyManager,
):
    backoff = 1.0
    max_backoff = 16.0
    max_retries = len(api_key_manager.keys) * 3
    retry_count = 0

    async with semaphore:
        while retry_count < max_retries:
            current_key = await api_key_manager.get_current_key()
            current_idx = await api_key_manager.get_current_index()

            headers = {
                "X-API-KEY": current_key,
                "Content-Type": "application/json",
            }
            payload = {
                "q": query,
                "location": LOCATION,
                "gl": GL,
                "hl": HL,
                "num": RESULTS_PER_PAGE,
                "page": page,
            }

            try:
                async with session.post(
                    BASE_URL, json=payload, headers=headers, timeout=30
                ) as r:
                    if r.status == 200:
                        data = await r.json()
                        organic_count = len(data.get("organic", []))
                        print(f"\n'{query}' стор.{page}: {organic_count} результатів")
                        return data

                    if r.status in (403, 429):
                        try:
                            err = await r.json()
                            msg = err.get("message", str(err))[:200]
                        except Exception:
                            msg = (await r.text())[:200]
                        print(f"\nHTTP {r.status} (ключ #{current_idx+1}): {msg}")
                        await api_key_manager.mark_key_failed()
                        if await api_key_manager.rotate_key():
                            await asyncio.sleep(backoff)
                            backoff = min(max_backoff, backoff * 2)
                            retry_count += 1
                            continue

                    if r.status >= 500:
                        print(f"\nСерверна помилка {r.status} — повтор через {backoff}с")
                        await asyncio.sleep(backoff)
                        backoff = min(max_backoff, backoff * 2)
                        retry_count += 1
                        continue

                    print(f"\nHTTP {r.status}")
                    return None

            except asyncio.TimeoutError:
                print(f"\nТаймаут '{query}'")
                await asyncio.sleep(backoff)
                backoff = min(max_backoff, backoff * 2)
                retry_count += 1
            except Exception as e:
                print(f"\nПомилка '{query}': {e}")
                await asyncio.sleep(backoff)
                backoff = min(max_backoff, backoff * 2)
                retry_count += 1

        print(f"\nПеревищено спроби для '{query}'")
        return None


async def process_keyword(
    kw: str,
    session: aiohttp.ClientSession,
    semaphore: asyncio.Semaphore,
    api_key_manager: APIKeyManager,
):
    """
    Парсимо keyword до MAX_POSITIONS (Top N),
    розкидаємо по сторінках 1..PAGES (де PAGES = ceil(MAX_POSITIONS / RESULTS_PER_PAGE))
    """
    tasks = []
    for page in range(1, PAGES + 1):
        tasks.append(
            serper_search_async(kw, page, session, semaphore, api_key_manager)
        )

    results = await asyncio.gather(*tasks, return_exceptions=True)
    keyword_data = []
    target_found = []

    for page, data in enumerate(results, 1):
        start_pos = (page - 1) * RESULTS_PER_PAGE + 1
        if data is None or isinstance(data, Exception):
            continue
        if isinstance(data, dict) and ("error" in data or "message" in data):
            continue

        items = data.get("organic", [])
        for idx, item in enumerate(items, start=start_pos):
            if idx > MAX_POSITIONS:
                break

            link = item.get("link", "")
            if not link or not link.startswith("http"):
                continue

            full_domain = get_full_domain(link)
            is_target = is_target_domain(full_domain)

            title = item.get("title", "") or ""
            snippet = item.get("snippet", "") or ""

            keyword_data.append(
                {
                    "Keyword": kw,
                    "Position": idx,
                    "Domain": full_domain,
                    "Title": title,
                    "Snippet": snippet,
                    "URL": link,
                    "Is_Target": is_target,
                    "Bucket": bucket_for_position(idx),
                }
            )

            if is_target:
                target_found.append(f"{full_domain} @{idx}")

    if target_found:
        print(
            f"\n'{kw}': знайдено {len(target_found)} таргет(ів): "
            f"{', '.join(target_found)}"
        )

    return keyword_data


# =========================
# Допоміжні функції
# =========================
def style_header(ws, headers):
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(
            start_color="1F4E78", end_color="1F4E78", fill_type="solid"
        )
        cell.alignment = Alignment(horizontal="center", vertical="center")


def autosize_columns(ws, max_width=100):
    for col in ws.columns:
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(
            max_len + 3, max_width
        )


def load_history():
    try:
        with open(HISTORY_FILE, "r", encoding="utf-8") as f:
            return json.load(f)
    except FileNotFoundError:
        return []
    except Exception as e:
        print(f"Помилка завантаження історії: {e}")
        return []


def save_history(all_rows, timestamp):
    history = load_history()
    history.append(
        {
            "timestamp": timestamp,
            "results": [
                {
                    "Keyword": r["Keyword"],
                    "Position": r["Position"],
                    "Domain": r["Domain"],
                    "Title": r["Title"],
                    "Snippet": r["Snippet"],
                    "URL": r["URL"],
                    "Is_Target": r["Is_Target"],
                }
                for r in all_rows
            ],
        }
    )
    MAX_HISTORY_ENTRIES = 10
    if len(history) > MAX_HISTORY_ENTRIES:
        history = history[-MAX_HISTORY_ENTRIES:]
    try:
        with open(HISTORY_FILE, "w", encoding="utf-8") as f:
            json.dump(history, f, ensure_ascii=False, indent=2)
        print(f"Історія збережена ({len(history)} записів)")
    except Exception as e:
        print(f"Не вдалося зберегти історію: {e}")


def get_position_history(history, keyword, domain):
    pos_hist = []
    for entry in history:
        ts = entry.get("timestamp", "")
        found = False
        for item in entry.get("results", []):
            if item["Keyword"] == keyword and item["Domain"] == domain:
                pos_hist.append({"timestamp": ts, "position": item["Position"]})
                found = True
                break
        if not found:
            pos_hist.append({"timestamp": ts, "position": None})
    return pos_hist


def calculate_trend(hist):
    valid = [p["position"] for p in hist if p["position"] is not None]
    if len(valid) <= 1:
        return "New" if valid else "No data"
    cur, prev = valid[-1], valid[-2]
    if cur < prev:
        return f"Up {prev - cur}"
    if cur > prev:
        return f"Down {cur - prev}"
    return "="


def get_average_position(hist):
    valid = [p["position"] for p in hist if p["position"] is not None]
    return round(sum(valid) / len(valid), 1) if valid else None


def get_best_position(hist):
    valid = [p["position"] for p in hist if p["position"] is not None]
    return min(valid) if valid else None


def get_worst_position(hist):
    valid = [p["position"] for p in hist if p["position"] is not None]
    return max(valid) if valid else None


# =========================
# Головна функція проєкту
# =========================
async def run_project(project_config: dict, progress_callback=None) -> str:
    """
    project_config:
      {
        "name": "...",
        "location": "...",
        "gl": "...",
        "hl": "...",
        "api_keys": [...],
        "target_domains": [...],
        "keywords": [...],
        "pages": 5,                 <-- ✅ тепер підтримується
        "max_positions": 30,        <-- або як раніше
        "history_file": "...json",
        "output_prefix": "..."
      }
    """
    global LOCATION, GL, HL, TARGET_DOMAINS, PAGES, MAX_POSITIONS, HISTORY_FILE, OUTPUT_FILE

    LOCATION = project_config["location"]
    GL = project_config["gl"]
    HL = project_config["hl"]
    TARGET_DOMAINS = set(project_config["target_domains"])

    # ✅ ГОЛОВНИЙ ФІКС: pages має пріоритет над max_positions
    pages_cfg = project_config.get("pages")
    maxpos_cfg = project_config.get("max_positions")

    if pages_cfg is not None:
        try:
            PAGES = max(1, int(pages_cfg))
        except Exception:
            PAGES = 3
        MAX_POSITIONS = PAGES * RESULTS_PER_PAGE
    else:
        MAX_POSITIONS = int(maxpos_cfg or 30)
        PAGES = max(1, math.ceil(MAX_POSITIONS / RESULTS_PER_PAGE))

    HISTORY_FILE = project_config["history_file"]

    # Будуємо динамічні бакети під MAX_POSITIONS
    build_buckets(MAX_POSITIONS)

    timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    timestamp_slug = datetime.datetime.now().strftime("%Y%m%d_%H%M")
    OUTPUT_FILE = f'{project_config["output_prefix"]}_{timestamp_slug}.xlsx'

    keywords = [k.strip() for k in project_config["keywords"] if k.strip()]

    print("-" * 90)
    print(
        f"Запуск: Serper.dev | ТОП-{MAX_POSITIONS} | Гео: {LOCATION} "
        f"+ СУБДОМЕНИ + ПОВНА ІСТОРІЯ"
    )
    print(f"{timestamp}")
    print(
        f"Ключів API: {len(project_config['api_keys'])} "
        f"| Таргет-доменів: {len(TARGET_DOMAINS)}"
    )
    print(f"Ключових слів: {len(keywords)} | Вивід: {OUTPUT_FILE}")
    print(f"Сторінок на ключ: {PAGES} (по {RESULTS_PER_PAGE} результатів)")
    print(f"Бакети: {', '.join(BUCKET_LABELS)}")
    print("-" * 90)

    all_rows = []
    domain_interval_counts = defaultdict(lambda: {label: 0 for label in BUCKET_LABELS})
    domain_keywords = defaultdict(list)

    api_key_manager = APIKeyManager(project_config["api_keys"])

    connector = aiohttp.TCPConnector(limit_per_host=MAX_CONCURRENT_REQUESTS, ssl=False)
    async with aiohttp.ClientSession(connector=connector) as session:
        semaphore = asyncio.Semaphore(MAX_CONCURRENT_REQUESTS)
        total_kw = len(keywords)
        for i, kw in enumerate(keywords, 1):
            keyword_data = await process_keyword(kw, session, semaphore, api_key_manager)
            all_rows.extend(keyword_data)

            for item in keyword_data:
                d = item["Domain"]
                b = item["Bucket"]
                domain_interval_counts[d][b] += 1
                domain_keywords[d].append({"keyword": item["Keyword"], "position": item["Position"]})

            if progress_callback is not None:
                progress_callback(i, total_kw, len(all_rows))

            print(
                f"\rОброблено: {i}/{total_kw} | "
                f"Знайдено позицій: {len(all_rows)}",
                end="",
                flush=True,
            )
            await asyncio.sleep(random.uniform(0.4, 0.9))

    print("\n")
    history = load_history()
    save_history(all_rows, timestamp)
    target_rows = [r for r in all_rows if r["Is_Target"]]

    # =========================
    # Excel
    # =========================
    wb = Workbook()

    # 1. Results
    ws_res = wb.active
    ws_res.title = "Results"
    headers_res = [
        "Project",
        "Keyword",
        "Position",
        "Domain",
        "Title",
        "Snippet",
        "URL",
        "Is_Target",
    ]
    style_header(ws_res, headers_res)
    for row in all_rows:
        ws_res.append(
            [
                project_config["name"],
                row["Keyword"],
                row["Position"],
                row["Domain"],
                row["Title"],
                row["Snippet"],
                row["URL"],
                "Yes" if row["Is_Target"] else "",
            ]
        )
        if row["Is_Target"]:
            for c in range(1, 9):
                ws_res.cell(row=ws_res.max_row, column=c).fill = PatternFill("solid", "C6EFCE")

    # 2. Target Domains Stats
    ws_target = wb.create_sheet("Target Domains Stats")
    headers_target = ["Domain", "Total"] + BUCKET_LABELS + ["Score", "Keywords"]
    style_header(ws_target, headers_target)

    target_data = []
    for domain in set(r["Domain"] for r in target_rows):
        stats = domain_interval_counts[domain]
        total = sum(stats.values())
        kw_list = "; ".join(sorted(set(item["keyword"] for item in domain_keywords[domain])))
        score = calculate_success_score(stats)
        row = [domain, total] + [stats[label] for label in BUCKET_LABELS] + [score, kw_list]
        target_data.append(row)

    score_index = 2 + len(BUCKET_LABELS)
    target_data.sort(key=lambda x: x[score_index], reverse=True)

    for row in target_data:
        ws_target.append(row)
        for c in range(1, len(headers_target) + 1):
            ws_target.cell(row=ws_target.max_row, column=c).fill = PatternFill("solid", "C6EFCE")

    # 3. Position Buckets
    ws_pos = wb.create_sheet("Position Buckets")
    headers_pos = ["Domain", "Total"] + BUCKET_LABELS + ["Score"]
    style_header(ws_pos, headers_pos)

    all_domains_rows = []
    for domain, stats in domain_interval_counts.items():
        total = sum(stats.values())
        if total == 0:
            continue
        score = calculate_success_score(stats)
        row_tuple = (domain, total, *[stats[label] for label in BUCKET_LABELS], score)
        all_domains_rows.append(row_tuple)

    all_domains_rows.sort(key=lambda x: x[2 + len(BUCKET_LABELS)], reverse=True)
    for row in all_domains_rows:
        ws_pos.append(row)
        if is_target_domain(row[0]):
            for c in range(1, len(headers_pos) + 1):
                ws_pos.cell(row=ws_pos.max_row, column=c).fill = PatternFill("solid", "C6EFCE")

    # 4. Dynamics (All Keywords)
    ws_dyn = wb.create_sheet("Dynamics (All Keywords)")
    prev = len(history)
    headers_dyn = [
        "Domain",
        "Keyword",
        "Status",
        "Current",
        "Trend",
    ] + [f"Parse {i}" for i in range(prev, 0, -1)] + [
        "Avg",
        "Best",
        "Worst",
        "URL",
        "Title",
    ]
    style_header(ws_dyn, headers_dyn)

    all_domain_keyword_pairs = set()
    for entry in history:
        for item in entry.get("results", []):
            if item.get("Is_Target"):
                all_domain_keyword_pairs.add((item["Domain"], item["Keyword"]))

    current_pairs = {}
    for row in target_rows:
        all_domain_keyword_pairs.add((row["Domain"], row["Keyword"]))
        current_pairs[(row["Domain"], row["Keyword"])] = row

    sorted_pairs = sorted(all_domain_keyword_pairs)

    for domain, keyword in sorted_pairs:
        hist = get_position_history(history, keyword, domain)

        has_historical_position = any(p["position"] is not None for p in hist)
        is_currently_present = (domain, keyword) in current_pairs

        current_pos = None
        current_url = ""
        current_title = ""
        is_lost = False

        if is_currently_present:
            current_pos = current_pairs[(domain, keyword)]["Position"]
            current_url = current_pairs[(domain, keyword)]["URL"]
            current_title = current_pairs[(domain, keyword)]["Title"]
        elif has_historical_position:
            last_idx = max(i for i, p in enumerate(hist) if p["position"] is not None)
            runs_since_seen = prev - last_idx

            if runs_since_seen <= 2:
                is_lost = True
                current_pos = "LOST"
            else:
                continue

        trend = calculate_trend(hist)
        positions = [p["position"] if p["position"] is not None else "—" for p in hist]
        avg = get_average_position(hist)
        best = get_best_position(hist)
        worst = get_worst_position(hist)

        status = (
            "Active"
            if (current_pos and current_pos != "LOST")
            else ("LOST" if is_lost else "No data")
        )

        row_data = [
            domain,
            keyword,
            status,
            current_pos if current_pos else "—",
            trend,
        ] + positions + [
            avg or "—",
            best or "—",
            worst or "—",
            current_url,
            current_title,
        ]

        ws_dyn.append(row_data)

        if is_lost or status == "LOST":
            for c in range(1, len(headers_dyn) + 1):
                cell = ws_dyn.cell(row=ws_dyn.max_row, column=c)
                cell.fill = PatternFill("solid", "FFB6C1")
                if c == 3:
                    cell.font = Font(color="FF0000", bold=True)
        else:
            for c in range(1, len(headers_dyn) + 1):
                cell = ws_dyn.cell(row=ws_dyn.max_row, column=c)
                cell.fill = PatternFill("solid", "C6EFCE")

        cell_trend = ws_dyn.cell(row=ws_dyn.max_row, column=5)
        if "Up" in str(cell_trend.value):
            cell_trend.font = Font(color="008000", bold=True)
        if "Down" in str(cell_trend.value):
            cell_trend.font = Font(color="FF0000", bold=True)

    # 5. Lost Keywords
    ws_lost = wb.create_sheet("Lost Keywords")
    headers_lost = [
        "Domain",
        "Keyword",
        "Last Seen Position",
        "Last Seen Date",
        "Days Since Lost",
    ]
    style_header(ws_lost, headers_lost)

    for domain, keyword in sorted_pairs:
        hist = get_position_history(history, keyword, domain)

        has_historical_position = any(p["position"] is not None for p in hist)
        is_currently_present = (domain, keyword) in current_pairs

        if has_historical_position and not is_currently_present:
            last_position = None
            last_date = None
            last_idx = None
            for i, p in reversed(list(enumerate(hist))):
                if p["position"] is not None:
                    last_position = p["position"]
                    last_date = p["timestamp"]
                    last_idx = i
                    break

            if last_idx is None:
                continue

            runs_since_seen = prev - last_idx
            if runs_since_seen > 2:
                continue

            days_lost = "—"
            if last_date:
                try:
                    last_dt = datetime.datetime.strptime(last_date, "%Y-%m-%d %H:%M:%S")
                    current_dt = datetime.datetime.strptime(timestamp, "%Y-%m-%d %H:%M:%S")
                    days_lost = (current_dt - last_dt).days
                except Exception:
                    pass

            ws_lost.append([domain, keyword, last_position or "—", last_date or "—", days_lost])
            for c in range(1, 6):
                ws_lost.cell(row=ws_lost.max_row, column=c).fill = PatternFill("solid", "FFB6C1")
                ws_lost.cell(row=ws_lost.max_row, column=c).font = Font(color="8B0000")

    # 6. History Summary
    ws_hist = wb.create_sheet("History Summary")
    headers_hist = ["Date", "Total Found", "Avg Pos"] + BUCKET_LABELS
    style_header(ws_hist, headers_hist)

    full_hist = history + [{"timestamp": timestamp, "results": all_rows}]
    for entry in full_hist:
        targets = [r for r in entry.get("results", []) if r.get("Is_Target")]
        if not targets:
            continue
        pos = [r["Position"] for r in targets]
        avg_pos = round(sum(pos) / len(pos), 1) if pos else 0

        row = [entry.get("timestamp", "—"), len(targets), avg_pos]
        for (start, end) in BUCKET_RANGES:
            row.append(sum(1 for p in pos if start <= p <= end))
        ws_hist.append(row)

    for ws in wb.worksheets:
        autosize_columns(ws, max_width=100)

    try:
        Path(OUTPUT_FILE).parent.mkdir(parents=True, exist_ok=True)
        wb.save(OUTPUT_FILE)
        print(f"\nГОТОВО! Файл збережено: {OUTPUT_FILE}")
        print(
            "Аркуші: Results • Target Domains Stats • Position Buckets • "
            "Dynamics (All Keywords) • Lost Keywords • History Summary"
        )
    except Exception as e:
        print(f"Помилка збереження: {e}")

    return OUTPUT_FILE
