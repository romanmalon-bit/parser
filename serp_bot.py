import asyncio
import json
import logging
import os
import signal
import fcntl
from pathlib import Path
from datetime import datetime, time
from typing import List, Optional, Dict, Tuple

from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application,
    CommandHandler,
    CallbackQueryHandler,
    ContextTypes,
    ConversationHandler,
    MessageHandler,
    filters,
)
from openpyxl import load_workbook

from parser_core import run_project  # НЕ чіпаємо parser_core.py

# =========================
# НАЛАШТУВАННЯ
# =========================
TELEGRAM_BOT_TOKEN = "8146349890:AAGvkkJnglQfQak0yRxX3JMGZ3zzbKSU-Eo"
PROJECTS_FILE = "projects.json"
USERS_FILE = "users.txt"
ADMIN_FILE = "admin_chat_id.txt"
LAST_HISTORY_DIR = "last_history"
LOCK_FILE = "/tmp/telegram_bot.lock"

DEFAULT_ADMIN_CHAT_ID = 909587225
ADMIN_CHAT_ID = int(os.getenv("ADMIN_CHAT_ID", str(DEFAULT_ADMIN_CHAT_ID)))

logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# =========================
# ERROR HANDLER
# =========================
async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logger.exception("Непіймана помилка:", exc_info=context.error)
    await send_error_to_admin(context, f"Непіймана помилка: {context.error}")

# =========================
# ADMIN & USERS
# =========================
def load_admin_chat_id() -> int:
    if ADMIN_CHAT_ID:
        return ADMIN_CHAT_ID
    try:
        if os.path.exists(ADMIN_FILE):
            return int(Path(ADMIN_FILE).read_text(encoding="utf-8").strip())
    except Exception:
        pass
    return 0

def save_admin_chat_id(chat_id: int):
    Path(ADMIN_FILE).write_text(str(chat_id), encoding="utf-8")

def load_users() -> set[int]:
    if not os.path.exists(USERS_FILE):
        return set()
    try:
        return {int(line.strip()) for line in Path(USERS_FILE).read_text(encoding="utf-8").splitlines() if line.strip()}
    except Exception:
        return set()

def save_users(users: set[int]):
    Path(USERS_FILE).write_text("\n".join(map(str, sorted(users))), encoding="utf-8")

def add_user(chat_id: int):
    users = load_users()
    if chat_id not in users:
        users.add(chat_id)
        save_users(users)

async def cmd_admin(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    save_admin_chat_id(chat_id)
    await update.message.reply_text(f"✅ ADMIN_CHAT_ID встановлено: {chat_id}")

async def cmd_users(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if update.effective_chat.id != load_admin_chat_id():
        await update.message.reply_text("❌ Доступ заборонено.")
        return
    users = load_users()
    text = f"👥 Активні користувачі ({len(users)}):\n" + "\n".join(map(str, sorted(users))) if users else "Немає активних користувачів."
    await update.message.reply_text(text)

async def send_error_to_admin(context: ContextTypes.DEFAULT_TYPE, error_text: str):
    admin_id = load_admin_chat_id()
    if not admin_id:
        return
    try:
        await context.bot.send_message(
            chat_id=admin_id,
            text=f"🚨 ПОМИЛКА:\n{error_text}\nЧас: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
        )
    except Exception as e:
        logger.error("Не вдалося надіслати помилку адміну: %s", e)

# =========================
# ПРОЄКТИ
# =========================
def load_projects() -> List[dict]:
    if not os.path.exists(PROJECTS_FILE):
        with open(PROJECTS_FILE, "w", encoding="utf-8") as f:
            json.dump({"projects": []}, f, ensure_ascii=False, indent=2)
        return []
    try:
        with open(PROJECTS_FILE, "r", encoding="utf-8") as f:
            return json.load(f).get("projects", [])
    except Exception:
        return []

def save_projects(projects: List[dict]):
    with open(PROJECTS_FILE, "w", encoding="utf-8") as f:
        json.dump({"projects": projects}, f, ensure_ascii=False, indent=2)

PROJECTS = load_projects()
PROJECTS_BY_NAME = {p["name"]: p for p in PROJECTS}

def reload_projects():
    global PROJECTS, PROJECTS_BY_NAME
    PROJECTS = load_projects()
    PROJECTS_BY_NAME = {p["name"]: p for p in PROJECTS}

# =========================
# STATE
# =========================
def get_state(context: ContextTypes.DEFAULT_TYPE):
    if "state" not in context.user_data:
        context.user_data["state"] = {"pages": 3, "projects": []}
    return context.user_data["state"]

# =========================
# SAFE SEND
# =========================
async def _safe_send_message(bot, chat_id: int, text: str) -> bool:
    try:
        await bot.send_message(chat_id=chat_id, text=text, parse_mode="Markdown")
        return True
    except Exception as e:
        logger.warning("send_message failed for %s: %s", chat_id, e)
        return False

async def _safe_send_document(bot, chat_id: int, path: Path, caption: str) -> bool:
    try:
        with path.open("rb") as f:
            await bot.send_document(chat_id=chat_id, document=f, caption=caption)
        return True
    except Exception as e:
        logger.warning("send_document failed for %s: %s", chat_id, e)
        return False

# =========================
# XLSX HELPERS
# =========================
def find_latest_xlsx(since_ts: float) -> Optional[Path]:
    latest = None
    latest_mtime = 0.0
    for p in Path(".").rglob("*.xlsx"):
        try:
            m = p.stat().st_mtime
            if m >= since_ts and m >= latest_mtime:
                latest = p
                latest_mtime = m
        except Exception:
            continue
    return latest

def find_previous_report(output_prefix: str, current_path: Path) -> Optional[Path]:
    candidates = [p for p in Path(".").rglob(f"{output_prefix}_*.xlsx") if p.resolve() != current_path.resolve()]
    if not candidates:
        return None
    candidates.sort(key=lambda x: x.stat().st_mtime, reverse=True)
    return candidates[0]

def read_target_domain_stats(xlsx_path: Path) -> Dict[str, float]:
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    if "Target Domains Stats" not in wb.sheetnames:
        return {}
    ws = wb["Target Domains Stats"]
    header = [str(c.value).strip() if c.value else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    idx = {name: i for i, name in enumerate(header)}
    domain_i = idx.get("Domain")
    kw_i = idx.get("Keywords")
    total_i = idx.get("Total")
    if domain_i is None:
        return {}
    out: Dict[str, float] = {}
    for row in ws.iter_rows(min_row=2):
        domain = row[domain_i].value
        if not domain:
            continue
        domain = str(domain).strip().lower()
        kw_count = 0
        if kw_i is not None:
            cell = row[kw_i].value
            if cell:
                kws = [k.strip() for k in str(cell).split(";") if k.strip()]
                kw_count = len(set(kws))
        if kw_count == 0 and total_i is not None:
            try:
                kw_count = int(row[total_i].value or 0)
            except Exception:
                pass
        out[domain] = float(kw_count)
    return out

def _badge(prev: float, now: float) -> str:
    if prev == 0 and now > 0:
        return "🟢"
    if prev > 0 and now == 0:
        return "🟥"
    if now > prev:
        return "🟢"
    if now < prev:
        if now * 2 < prev:
            return "🟥"
        return "🔻"
    return "⚪"

def format_delta_report(prev_map: Dict[str, float], cur_map: Dict[str, float], top_n: int = 30) -> str:
    domains = sorted(set(prev_map.keys()) | set(cur_map.keys()))
    
    rows: List[Tuple[float, float, str, float, float]] = []
    summary = {"kw_up": 0, "kw_down": 0, "kw_severe": 0, "kw_new": 0, "kw_lost": 0, "kw_same": 0}

    for d in domains:
        pkw = prev_map.get(d, 0.0)
        nkw = cur_map.get(d, 0.0)
        
        if pkw == 0 and nkw > 0:
            summary["kw_new"] += 1
            summary["kw_up"] += 1
        elif pkw > 0 and nkw == 0:
            summary["kw_lost"] += 1
            summary["kw_down"] += 1
            summary["kw_severe"] += 1
        elif nkw > pkw:
            summary["kw_up"] += 1
        elif nkw < pkw:
            summary["kw_down"] += 1
            if nkw * 2 < pkw:
                summary["kw_severe"] += 1
        else:
            summary["kw_same"] += 1
        
        rows.append((abs(nkw - pkw), nkw, d, pkw, nkw))

    rows.sort(key=lambda x: (-x[0], -x[1]))
    rows = rows[:top_n]

    lines = []
    lines.append(
        f"📊 *Динаміка Keywords vs попередній парсинг* (топ {len(rows)} доменів)\n"
        f"🟢 Зростання: {summary['kw_up']}  🔻 Падіння: {summary['kw_down']} (🟥 сильне: {summary['kw_severe']})\n"
        f"NEW: {summary['kw_new']}  LOST: {summary['kw_lost']}  ⚪ Без змін: {summary['kw_same']}\n"
    )
    lines.append("```text")
    lines.append(f"{'Бадж':<2} {'Prev → Now':^11} {'ΔKW':>6} {'Domain'}")
    lines.append("───┼─────────────┼──────┼───────────────────────────────────")

    for _, _, d, pkw, nkw in rows:
        badge = _badge(pkw, nkw)
        dkw = int(nkw - pkw)
        delta_str = f"{dkw:+}".rjust(5)
        prev_now = f"{int(pkw):>4} → {int(nkw):<4}"
        domain = d[:35]
        lines.append(f"{badge:<2} {prev_now} {delta_str} {domain}")

    if not rows:
        lines.append("  Немає даних для порівняння.")
    lines.append("```")
    return "\n".join(lines)

def add_history_sheet_if_needed(xlsx_path: Path, project_name: str):
    Path(LAST_HISTORY_DIR).mkdir(exist_ok=True)
    history_path = Path(LAST_HISTORY_DIR) / f"{project_name}.json"
    try:
        wb = load_workbook(xlsx_path, read_only=False)
        data_to_save = {}
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            data = [[cell.value for cell in row] for row in ws.iter_rows()]
            data_to_save[sheet_name] = data
        history_path.write_text(json.dumps(data_to_save, ensure_ascii=False, indent=2), encoding="utf-8")

        prev_reports = list(Path(".").rglob(f"*_{project_name}_*.xlsx"))
        if len(prev_reports) > 1 or (history_path.exists() and history_path.stat().st_size > 10):
            if "History" in wb.sheetnames:
                wb.remove(wb["History"])
            history_ws = wb.create_sheet("History")
            if history_path.exists():
                old_data = json.loads(history_path.read_text(encoding="utf-8"))
                for sheet_name, rows in old_data.items():
                    history_ws.append([f"=== Попередній парсинг: {sheet_name} ==="])
                    for row in rows:
                        history_ws.append(row)
                    history_ws.append([])
        wb.save(xlsx_path)
    except Exception as e:
        logger.error(f"Помилка додавання історії для {project_name}: {e}")

def cleanup_old_reports(output_prefix: str, keep_last: int = 2):
    try:
        files = list(Path(".").rglob(f"{output_prefix}_*.xlsx"))
        if len(files) <= keep_last:
            return
        files.sort(key=lambda x: x.stat().st_mtime, reverse=True)
        for old_file in files[keep_last:]:
            try:
                old_file.unlink()
                logger.info(f"Видалено старий файл: {old_file.name}")
            except Exception as e:
                logger.warning(f"Не вдалося видалити {old_file.name}: {e}")
    except Exception as e:
        logger.error(f"Помилка при очищенні файлів: {e}")

# =========================
# КЛАВІАТУРИ
# =========================
def kb_main(st):
    return InlineKeyboardMarkup([
        [InlineKeyboardButton("🧩 Виберіть проєкти", callback_data="projects")],
        [InlineKeyboardButton(f"📄 Сторінки: {st['pages']} (топ {st['pages']*10})", callback_data="pages")],
        [InlineKeyboardButton("▶️ Запустити парсинг", callback_data="run")],
        [InlineKeyboardButton("➕ Додати новий проєкт", callback_data="add_project")],
        [InlineKeyboardButton("🗑 Видалити проєкт", callback_data="delete")],
        [InlineKeyboardButton("ℹ️ Довідка", callback_data="info")],
    ])

def kb_projects(st):
    buttons = [[InlineKeyboardButton(f"{'✅' if p['name'] in st['projects'] else '☑️'} {p['name']}", callback_data=f"toggle:{p['name']}")] for p in PROJECTS]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back")])
    return InlineKeyboardMarkup(buttons)

def kb_pages():
    return InlineKeyboardMarkup([
        [InlineKeyboardButton(str(i), callback_data=f"setpages:{i}") for i in range(1, 6)],
        [InlineKeyboardButton(str(i), callback_data=f"setpages:{i}") for i in range(6, 11)],
        [InlineKeyboardButton("⬅️ Назад", callback_data="back")],
    ])

def kb_delete():
    buttons = [[InlineKeyboardButton(f"🗑 {p['name']}", callback_data=f"del:{p['name']}")] for p in PROJECTS]
    buttons.append([InlineKeyboardButton("⬅️ Назад", callback_data="back")])
    return InlineKeyboardMarkup(buttons)

# =========================
# START & CALLBACK
# =========================
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    add_user(chat_id)
    st = get_state(context)
    await update.effective_message.reply_text(
        "Привіт! Це бот для парсингу SERP.\n"
        "— Ручний парсинг: виберіть проєкти + сторінки → ▶️\n"
        "— Автопарсинг: о 07:00, 12:00, 17:00 за Києвом\n\n"
        "Оберіть дію:",
        reply_markup=kb_main(st)
    )

async def callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    await query.answer()
    st = get_state(context)
    data = query.data
    chat_id = query.message.chat_id

    if data == "projects":
        reload_projects()
        await query.edit_message_text("Виберіть проєкти:", reply_markup=kb_projects(st))
    elif data.startswith("toggle:"):
        name = data.split(":", 1)[1]
        if name in st["projects"]:
            st["projects"].remove(name)
        else:
            st["projects"].append(name)
        await query.edit_message_reply_markup(reply_markup=kb_projects(st))
    elif data == "pages":
        await query.edit_message_text("Кількість сторінок:", reply_markup=kb_pages())
    elif data.startswith("setpages:"):
        st["pages"] = int(data.split(":")[1])
        await query.edit_message_text("Оновлено.", reply_markup=kb_main(st))
    elif data == "run":
        if not st["projects"]:
            await query.edit_message_text("Оберіть хоча б один проєкт.", reply_markup=kb_main(st))
            return
        pages = st["pages"]
        top_n = pages * 10
        await query.edit_message_text(f"⏳ Ручний парсинг: {len(st['projects'])} проєктів, TOP {top_n}", reply_markup=kb_main(st))

        async def runner():
            try:
                for i, name in enumerate(st["projects"], 1):
                    reload_projects()
                    project = PROJECTS_BY_NAME.get(name)
                    if not project:
                        await _safe_send_message(context.bot, chat_id, f"⚠️ Проєкт «{name}» не знайдено.")
                        continue
                    cfg = dict(project)
                    cfg["max_positions"] = top_n
                    output_prefix = cfg.get("output_prefix", "report")

                    await _safe_send_message(context.bot, chat_id, f"▶️ [{i}/{len(st['projects'])}] Парсю «{name}»")
                    start_ts = datetime.now().timestamp()

                    try:
                        out_path = await run_project(cfg)
                    except Exception as e:
                        await _safe_send_message(context.bot, chat_id, f"🚨 Помилка в «{name}»: {e}")
                        await send_error_to_admin(context, f"Помилка в «{name}»: {e}")
                        continue

                    xlsx_path = None
                    if isinstance(out_path, str):
                        p = Path(out_path)
                        if p.exists():
                            xlsx_path = p
                    if not xlsx_path:
                        xlsx_path = find_latest_xlsx(start_ts)

                    if xlsx_path and xlsx_path.exists():
                        add_history_sheet_if_needed(xlsx_path, name)
                        cleanup_old_reports(output_prefix)  # Залишаємо тільки 2 файли
                        await _safe_send_message(context.bot, chat_id, f"✅ «{name}» готово")
                        await _safe_send_document(context.bot, chat_id, xlsx_path, caption=xlsx_path.name)

                        prev_xlsx = find_previous_report(output_prefix, xlsx_path)
                        if prev_xlsx and prev_xlsx.exists():
                            prev_stats = read_target_domain_stats(prev_xlsx)
                            cur_stats = read_target_domain_stats(xlsx_path)
                            msg = format_delta_report(prev_stats, cur_stats)
                            await _safe_send_message(context.bot, chat_id, msg)
                        else:
                            await _safe_send_message(context.bot, chat_id, "ℹ️ Перший звіт — порівняння немає.")
                    else:
                        await _safe_send_message(context.bot, chat_id, "✅ Виконано, але файл не знайдено.")
                await _safe_send_message(context.bot, chat_id, "🏁 Ручний парсинг завершено.")
            except Exception as e:
                await send_error_to_admin(context, f"runner error: {e}")

        context.application.create_task(runner())

    elif data == "add_project":
        await query.edit_message_text("Виконай команду /addproject")
    elif data == "delete":
        reload_projects()
        await query.edit_message_text("Видалити проєкт:", reply_markup=kb_delete())
    elif data.startswith("del:"):
        name = data.split(":", 1)[1]
        projects = [p for p in load_projects() if p["name"] != name]
        save_projects(projects)
        reload_projects()
        if name in st["projects"]:
            st["projects"].remove(name)
        await query.edit_message_text(f"Проєкт «{name}» видалено.", reply_markup=kb_main(st))
    elif data == "info":
        await query.edit_message_text(
            "ℹ️ /start — меню\n/addproject — додати проєкт\n/cancel — скасувати\n/admin — алерти",
            reply_markup=kb_main(st)
        )
    elif data == "back":
        await query.edit_message_text("Меню:", reply_markup=kb_main(st))

# =========================
# ДОДАВАННЯ ПРОЄКТУ
# =========================
(NAME, LOCATION, LANGUAGE, API_KEYS, TARGET_DOMAINS, KEYWORDS, OUTPUT_PREFIX, HISTORY_FILE) = range(8)

async def start_add_project(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Крок 1: Назва проєкту")
    context.user_data["new_project"] = {}
    return NAME

async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE):
    reload_projects()
    name = update.message.text.strip()
    if name in PROJECTS_BY_NAME:
        await update.message.reply_text("Назва зайнята. Інша?")
        return NAME
    context.user_data["new_project"]["name"] = name
    await update.message.reply_text("Крок 2: Країна (location, напр. France)")
    return LOCATION

async def get_location(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_project"]["location"] = update.message.text.strip()
    await update.message.reply_text("Крок 3: Код мови (hl/gl, напр. fr)")
    return LANGUAGE

async def get_language(update: Update, context: ContextTypes.DEFAULT_TYPE):
    lang = update.message.text.strip()
    context.user_data["new_project"]["hl"] = lang
    context.user_data["new_project"]["gl"] = lang
    await update.message.reply_text("Крок 4: API ключі (через кому)")
    return API_KEYS

async def get_api_keys(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keys = [k.strip() for k in update.message.text.split(",") if k.strip()]
    context.user_data["new_project"]["api_keys"] = keys
    await update.message.reply_text("Крок 5: Таргет-домени (по одному або через кому)")
    return TARGET_DOMAINS

async def get_target_domains(update: Update, context: ContextTypes.DEFAULT_TYPE):
    domains = [d.strip() for d in update.message.text.replace(",", "\n").split("\n") if d.strip()]
    context.user_data["new_project"]["target_domains"] = domains
    await update.message.reply_text("Крок 6: Ключові слова (по одному або через кому)")
    return KEYWORDS

async def get_keywords(update: Update, context: ContextTypes.DEFAULT_TYPE):
    keywords = [k.strip() for k in update.message.text.replace(",", "\n").split("\n") if k.strip()]
    context.user_data["new_project"]["keywords"] = keywords
    await update.message.reply_text("Крок 7: Префікс файлу (напр. serp_fr)")
    return OUTPUT_PREFIX

async def get_output_prefix(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_project"]["output_prefix"] = update.message.text.strip()
    await update.message.reply_text("Крок 8: Ім'я файлу історії (будь-яке)")
    return HISTORY_FILE

async def get_history_file(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data["new_project"]["history_file"] = update.message.text.strip()
    new_project = context.user_data["new_project"]
    projects = load_projects()
    projects.append(new_project)
    save_projects(projects)
    reload_projects()
    await update.message.reply_text(f"✅ Проєкт «{new_project['name']}» додано!", reply_markup=kb_main(get_state(context)))
    context.user_data.pop("new_project", None)
    return ConversationHandler.END

async def cancel_add_project(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("❌ Скасовано.", reply_markup=kb_main(get_state(context)))
    context.user_data.pop("new_project", None)
    return ConversationHandler.END

# =========================
# AUTO PARSING
# =========================
AUTO_LOCK = asyncio.Lock()

async def auto_parsing_task(context: ContextTypes.DEFAULT_TYPE):
    if AUTO_LOCK.locked():
        return
    async with AUTO_LOCK:
        users = load_users()
        if not users:
            return
        reload_projects()
        if not PROJECTS:
            for uid in users:
                await _safe_send_message(context.bot, uid, "⚠️ Немає проєктів")
            return

        for uid in users:
            await _safe_send_message(context.bot, uid, f"🤖 Автопарсинг стартував ({len(PROJECTS)} проєктів, TOP-30)")

        for i, project in enumerate(PROJECTS, 1):
            name = project.get("name", "Unnamed")
            cfg = dict(project)
            cfg["max_positions"] = 30
            output_prefix = cfg.get("output_prefix", "report")

            for uid in users:
                await _safe_send_message(context.bot, uid, f"▶️ [{i}/{len(PROJECTS)}] Парсю «{name}»")

            start_ts = datetime.now().timestamp()
            try:
                out_path = await run_project(cfg)
            except Exception as e:
                msg = f"🚨 Помилка в «{name}»: {e}"
                for uid in users:
                    await _safe_send_message(context.bot, uid, msg)
                await send_error_to_admin(context, msg)
                continue

            xlsx_path = None
            if isinstance(out_path, str):
                p = Path(out_path)
                if p.exists():
                    xlsx_path = p
            if not xlsx_path:
                xlsx_path = find_latest_xlsx(start_ts)

            if xlsx_path and xlsx_path.exists():
                add_history_sheet_if_needed(xlsx_path, name)
                cleanup_old_reports(output_prefix)  # Залишаємо тільки 2 файли
                for uid in users:
                    await _safe_send_message(context.bot, uid, f"✅ «{name}» готово")
                    await _safe_send_document(context.bot, uid, xlsx_path, caption=f"AUTO {xlsx_path.name}")

                    prev_xlsx = find_previous_report(output_prefix, xlsx_path)
                    if prev_xlsx and prev_xlsx.exists():
                        prev_stats = read_target_domain_stats(prev_xlsx)
                        cur_stats = read_target_domain_stats(xlsx_path)
                        msg = format_delta_report(prev_stats, cur_stats)
                        await _safe_send_message(context.bot, uid, msg)
                    else:
                        await _safe_send_message(context.bot, uid, "ℹ️ Перший автозвіт — порівняння немає.")
            else:
                for uid in users:
                    await _safe_send_message(context.bot, uid, f"✅ «{name}» виконано, файл не знайдено.")

        for uid in users:
            await _safe_send_message(context.bot, uid, "🏁 Автопарсинг завершено.")

# =========================
# MAIN
# =========================
def main():
    if not TELEGRAM_BOT_TOKEN:
        raise RuntimeError("TELEGRAM_BOT_TOKEN не заданий!")

    try:
        lock_fd = os.open(LOCK_FILE, os.O_CREAT | os.O_WRONLY)
        fcntl.flock(lock_fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except (OSError, IOError):
        logger.error("Інша інстанція бота вже працює. Зупиняємо цю.")
        return

    app = Application.builder().token(TELEGRAM_BOT_TOKEN).build()

    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("admin", cmd_admin))
    app.add_handler(CommandHandler("users", cmd_users))
    app.add_handler(CallbackQueryHandler(callback))

    conv = ConversationHandler(
        entry_points=[CommandHandler("addproject", start_add_project)],
        states={
            NAME: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            LOCATION: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_location)],
            LANGUAGE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_language)],
            API_KEYS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_api_keys)],
            TARGET_DOMAINS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_target_domains)],
            KEYWORDS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_keywords)],
            OUTPUT_PREFIX: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_output_prefix)],
            HISTORY_FILE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_history_file)],
        },
        fallbacks=[CommandHandler("cancel", cancel_add_project)],
    )
    app.add_handler(conv)
    app.add_error_handler(error_handler)

    # Автопарсинг о 07:00, 12:00, 17:00 за Києвом (взимку, UTC+2)
    app.job_queue.run_daily(auto_parsing_task, time=time(hour=5, minute=0))
    app.job_queue.run_daily(auto_parsing_task, time=time(hour=10, minute=0))
    app.job_queue.run_daily(auto_parsing_task, time=time(hour=15, minute=0))

    logger.info("Автопарсинг заплановано на 07:00, 12:00, 17:00 за київським часом")
    logger.info("Бот запущено і працює (polling активний)")

    async def stop_bot():
        logger.info("Зупиняємо бота...")
        await app.stop()
        await app.shutdown()
        os.close(lock_fd)
        try:
            os.unlink(LOCK_FILE)
        except OSError:
            pass
        logger.info("Бот зупинено")

    loop = asyncio.get_event_loop()
    for sig in (signal.SIGTERM, signal.SIGINT):
        loop.add_signal_handler(sig, lambda: asyncio.create_task(stop_bot()))

    try:
        app.run_polling(drop_pending_updates=True)
    finally:
        asyncio.run(stop_bot())

if __name__ == "__main__":
    main()
